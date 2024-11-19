import os
import glob
import pandas as pd
from commonroad.common.file_reader import CommonRoadFileReader
from commonroad.visualization.mp_renderer import MPRenderer
from commonroad_route_planner.utility.visualization import visualize_route
from commonroad_route_planner.route_planner import RoutePlanner
from commonroad_route_planner.reference_path_planner import ReferencePathPlanner
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

from commonroad.scenario.intersection import IntersectionIncomingElement

def process_scenario_files(scenarios_dir, output_dir, excel_file):
    # Create the output directory
    os.makedirs(output_dir, exist_ok=True)


    # Create Excel files
    wb = Workbook()
    ws = wb.active
    ws.title = "Scenario Images"
    ws.append(["Scenario ID", "MPRenderer Image", "Visualize Route Image"])

    # Set the width of columns B and C to 300
    ws.column_dimensions['A'].width = 22.45
    ws.column_dimensions['B'].width = 50  
    ws.column_dimensions['C'].width = 50

    # Iterate through the XML files in the scenarios directory
    for xml_file in glob.glob(os.path.join(scenarios_dir, "*.xml")):
        scenario_id = os.path.splitext(os.path.basename(xml_file))[0]

        # Reading scenarios and planning problem sets
        scenario, planning_problem_set = CommonRoadFileReader(xml_file).open()
        planning_problem = list(planning_problem_set.planning_problem_dict.values())[0]

        interections=scenario._lanelet_network._intersections

        # Renders scenes and planning issues using MPRenderer and generates PNG files
        renderer = MPRenderer(figsize=(12, 12))
        scenario.draw(renderer)
        planning_problem.draw(renderer)
        renderer.render()
        plt.margins(0, 0)
        mp_renderer_img_path = os.path.join(output_dir, f"{scenario_id}_mp_renderer.png")
        plt.savefig(mp_renderer_img_path)
        plt.close()

        route_planner = RoutePlanner(lanelet_network=scenario.lanelet_network,  planning_problem=planning_problem,scenario=scenario)
        routes= route_planner.plan_routes()

         # Instantiate reference path planner and plan reference path
        ref_path_planner: ReferencePathPlanner = ReferencePathPlanner(
            lanelet_network=scenario.lanelet_network,
            planning_problem=planning_problem,
            routes=routes,
        )

        reference_path = ref_path_planner.plan_shortest_reference_path(
            retrieve_shortest=True, consider_least_lance_changes=True
        )
        # Generating PNG files with the visualize_route function
        save_imgs = True
        visualize_route(
            reference_path=reference_path,
            scenario=scenario,
            planning_problem=planning_problem,
            save_img=save_imgs,
            draw_route_lanelets=True,
            draw_reference_path=True,
        )
        visualize_route_img_path = os.path.join(output_dir, f"{scenario_id}_visualize_route.png")
        os.makedirs(
            os.path.dirname(visualize_route_img_path), exist_ok=True
        )  # Ensure the directory exists
        plt.savefig(visualize_route_img_path, format="png")


        # Embedding the generated PNG file into an Excel file
        ws.append([scenario_id, "", ""])
        mp_renderer_img = Image(mp_renderer_img_path)
        visualize_route_img = Image(visualize_route_img_path)

        # Control image size
        mp_renderer_img.width, mp_renderer_img.height = resize_image(mp_renderer_img_path, 300)
        visualize_route_img.width, visualize_route_img.height = resize_image(visualize_route_img_path, 300)


        ws.add_image(mp_renderer_img, f"B{ws.max_row}")
        ws.add_image(visualize_route_img, f"C{ws.max_row}")

        # Set the height of the current row to 300
        ws.row_dimensions[ws.max_row].height = 225 



    # Save Excel file
    wb.save(excel_file)


    
 # Controls the size of the image, keeping the original proportions and setting the height to 300 pixels.
def resize_image(image_path, target_height):
    with PILImage.open(image_path) as img:
        width, height = img.size
        aspect_ratio = width / height
        new_width = int(target_height * aspect_ratio)
        return new_width, target_height


scenarios_dir = "../scenarios/inD_repair"
output_dir = "../output/pic"
excel_file = "../output/scenario_images.xlsx"
process_scenario_files(scenarios_dir, output_dir, excel_file)